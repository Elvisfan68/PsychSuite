"""
data_writer.py — Shared incremental Excel writer for PsychSuite.
Saves the workbook after every row so a crash loses at most one trial.
"""
import os
from datetime import datetime

try:
    import openpyxl
except ImportError as e:
    raise ImportError(
        "PsychSuite requires openpyxl. Install dependencies: pip install -r requirements.txt"
    ) from e


class IncrementalExcelWriter:
    """
    Opens (or creates) an Excel workbook and writes rows to a named sheet
    one at a time, saving to disk after every write.
    """

    def __init__(self, filepath: str, sheet_name: str):
        self.filepath = filepath
        self.sheet_name = sheet_name
        self.headers = None

        dirpath = os.path.dirname(filepath)
        if dirpath:
            os.makedirs(dirpath, exist_ok=True)

        if os.path.exists(filepath):
            self.wb = openpyxl.load_workbook(filepath)
            if sheet_name in self.wb.sheetnames:
                self.ws = self.wb[sheet_name]
                first_row = [cell.value for cell in self.ws[1]]
                if first_row and first_row[0] is not None:
                    self.headers = first_row
            else:
                self.ws = self.wb.create_sheet(sheet_name)
        else:
            self.wb = openpyxl.Workbook()
            self.ws = self.wb.active
            self.ws.title = sheet_name

        self.wb.save(self.filepath)

    def write_row(self, row_dict: dict):
        """Append one row dict and immediately save to disk."""
        if self.headers is None:
            self.headers = list(row_dict.keys())
            self.ws.append(self.headers)

        row = [row_dict.get(h, '') for h in self.headers]
        self.ws.append(row)
        self.wb.save(self.filepath)

    def log_session_event(self, message: str):
        """Append a row to _SessionLog (same file) for crash recovery / audit trail."""
        log_name = '_SessionLog'
        if log_name not in self.wb.sheetnames:
            lg = self.wb.create_sheet(log_name)
            lg.append(['Timestamp', 'Data_Sheet', 'Message'])
        else:
            lg = self.wb[log_name]
        lg.append([
            datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            self.sheet_name,
            message,
        ])
        self.wb.save(self.filepath)

    def close(self):
        self.wb.save(self.filepath)


def make_excel_path(data_dir: str, participant_id: str, treatment: str) -> str:
    """Build a timestamped .xlsx filepath."""
    ts = datetime.now().strftime('%Y%m%d_%H%M%S')
    pid = (participant_id or 'unknown').replace(' ', '_')
    tx = (treatment or 'no_tx').replace(' ', '_')
    return os.path.join(data_dir, f"{pid}_{tx}_{ts}.xlsx")
